#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Liebherr PO -> Excel  |  Masaüstü Uygulaması (CEKAPIM)
======================================================
Dosya seçme pencereli, tek dosyalık masaüstü uygulaması.
Tek kalemli veya çok kalemli PO'larla, tek veya birden çok PDF ile çalışır.

Çalıştırma (Python ile):
    python liebherr_po_gui.py

.exe yapmak için (Windows'ta):
    pip install pyinstaller pdfplumber openpyxl
    pyinstaller --onefile --windowed --name "Liebherr_PO_Excel" liebherr_po_gui.py
    -> dist/Liebherr_PO_Excel.exe
"""

import glob
import os
import re
import sys
import threading
from dataclasses import dataclass, field
from typing import List, Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================================
#  1) AYRIŞTIRMA ÇEKİRDEĞİ  (CLI sürümüyle aynı, test edilmiş mantık)
# ============================================================================

@dataclass
class LineItem:
    pos: str = ""
    part_no: str = ""
    drawing_index: str = ""
    lhb_no: str = ""
    diameter: Optional[float] = None
    length1: Optional[float] = None
    length2: Optional[float] = None
    coating: str = ""
    material_no: str = ""
    spec_index: List[str] = field(default_factory=list)
    qty: Optional[float] = None
    unit_price: Optional[float] = None
    delivery_date: str = ""
    raw_desc: str = ""


@dataclass
class Order:
    source_file: str = ""
    order_no: str = ""
    order_type: str = ""
    buyer: str = ""
    order_date: str = ""
    supplier_no: str = ""
    stated_total: Optional[float] = None
    items: List[LineItem] = field(default_factory=list)


def eu_to_float(s: str) -> Optional[float]:
    if s is None:
        return None
    s = s.strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def extract_text(pdf_path: str) -> str:
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


RE_ITEM_HEADER = re.compile(
    r"^(?P<pos>\d+)\s+(?P<part>\d+)\s+"
    r"(?P<date>\d{4}-\d{2}-\d{2})\s+"
    r"(?P<qty>[\d.,]+)\s+st\s+"
    r"(?P<price>[\d.,]+)\s*EUR\s+st\s*$"
)
RE_DESC = re.compile(r"^\((?P<seq>\d+)\)\s+(?P<desc>.+)$")
RE_PIN = re.compile(
    r"^PIN\s+"
    r"(?P<d>[\d.,]+)\s*[Xx]\s*(?P<l1>[\d.,]+)\s*[Xx]\s*(?P<l2>[\d.,]+)\s+"
    r"(?P<coat>VERCHR\.?|verchr\.?)?\s*"
    r"(?P<mat>.+?)"
    r"-(?P<idx>\d+)\s*$"
)
RE_SPEC = re.compile(r"^Specification\s*-\s*Index\s+(?P<val>.+?)\s*$")
RE_DRAW = re.compile(r"^Drawing\s*Index\s+(?P<val>\S+)")


def parse_order(pdf_path: str) -> Order:
    text = extract_text(pdf_path)
    lines = [ln.strip() for ln in text.splitlines()]
    order = Order(source_file=os.path.basename(pdf_path))

    if re.search(r"First samples|Erstmuster", text, re.I):
        order.order_type = "İlk Numune (Erstmuster)"
    elif re.search(r"Series|Serie", text):
        order.order_type = "Seri (Serie)"

    m = re.search(r"Buyer\s+Order-No\.\s*\n(.+?)\s+(\d{6,})", text)
    if m:
        order.buyer = m.group(1).strip()
        order.order_no = m.group(2).strip()

    m = re.search(r"Date\s+Supplier-No\.\s*\n(\d{4}-\d{2}-\d{2})\s+(\d+)", text)
    if m:
        order.order_date = m.group(1)
        order.supplier_no = m.group(2)

    m = re.search(r"Total\s*EUR\s*\n[\d.,]+\s+([\d.,]+)", text)
    if m:
        order.stated_total = eu_to_float(m.group(1))

    current: Optional[LineItem] = None
    for ln in lines:
        mh = RE_ITEM_HEADER.match(ln)
        if mh:
            current = LineItem(
                pos=mh.group("pos"),
                part_no=mh.group("part"),
                qty=eu_to_float(mh.group("qty")),
                unit_price=eu_to_float(mh.group("price")),
                delivery_date=mh.group("date"),
            )
            order.items.append(current)
            continue
        if current is None:
            continue
        md = RE_DESC.match(ln)
        if md:
            desc = md.group("desc").strip()
            current.raw_desc = desc
            mp = RE_PIN.match(desc)
            if mp:
                current.diameter = eu_to_float(mp.group("d"))
                current.length1 = eu_to_float(mp.group("l1"))
                current.length2 = eu_to_float(mp.group("l2"))
                if mp.group("coat"):
                    current.coating = "KROM"
                current.material_no = mp.group("mat").strip()
                current.drawing_index = mp.group("idx")
            continue
        ms = RE_SPEC.match(ln)
        if ms:
            current.spec_index.append(ms.group("val").strip())
            continue
        mdw = RE_DRAW.match(ln)
        if mdw:
            current.drawing_index = mdw.group("val").strip()
            continue

    for it in order.items:
        it.lhb_no = f"{it.part_no}-{it.drawing_index}" if it.drawing_index else it.part_no
    return order


# ============================================================================
#  2) EXCEL ÇIKTISI
# ============================================================================

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
MANUAL_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

COLUMNS = [
    ("Sipariş No", 14, CENTER, False),
    ("Sipariş Tipi", 22, LEFT, False),
    ("LHB No", 18, CENTER, False),
    ("Parça No", 13, CENTER, False),
    ("Çizim İnd.", 10, CENTER, False),
    ("PİM ÇAPI Ø (mm)", 15, CENTER, False),
    ("ÖLÇÜ-1 (mm)", 12, CENTER, False),
    ("ÖLÇÜ-2 (mm)", 12, CENTER, False),
    ("KAPLAMA", 12, CENTER, False),
    ("Malzeme / Spec İnd.", 26, LEFT, False),
    ("Adet", 9, CENTER, False),
    ("Birim Fiyat (EUR)", 15, CENTER, False),
    ("Toplam (EUR)", 14, CENTER, False),
    ("Teslim Tarihi", 13, CENTER, False),
    ("Sipariş Tarihi", 13, CENTER, False),
    ("Alıcı", 18, LEFT, False),
    ("KAYNAK ÖLÇÜSÜ", 14, CENTER, True),
    ("KAYNAK ÇAPI", 13, CENTER, True),
    ("FLANŞ KALINLIĞI", 15, CENTER, True),
    ("ONAY DURUMU", 14, CENTER, True),
    ("Açıklama (ham)", 42, LEFT, False),
]


def build_workbook(orders: List[Order], out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Siparişler"
    for c, (title, width, _a, _m) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    r = 2
    for order in orders:
        for it in order.items:
            spec = "; ".join(it.spec_index)
            material = it.material_no + (f"  |  {spec}" if spec else "")
            values = [
                order.order_no, order.order_type, it.lhb_no, it.part_no,
                it.drawing_index, it.diameter, it.length1, it.length2,
                it.coating, material, it.qty, it.unit_price, None,
                it.delivery_date, order.order_date, order.buyer,
                None, None, None, None, it.raw_desc,
            ]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = CELL_FONT
                cell.border = BORDER
                cell.alignment = COLUMNS[c - 1][2]
                if COLUMNS[c - 1][3]:
                    cell.fill = MANUAL_FILL
            ws.cell(row=r, column=13, value=f"=K{r}*L{r}")
            ws.cell(row=r, column=12).number_format = "#,##0.00"
            ws.cell(row=r, column=13).number_format = "#,##0.00"
            r += 1

    ws2 = wb.create_sheet("Özet")
    sum_cols = ["Kaynak Dosya", "Sipariş No", "Tip", "Alıcı", "Tarih",
                "Kalem Sayısı", "Hesaplanan Toplam (EUR)",
                "PDF'teki Toplam (EUR)", "Kontrol"]
    for c, title in enumerate(sum_cols, start=1):
        cell = ws2.cell(row=1, column=c, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for c, w in enumerate([34, 14, 22, 20, 13, 13, 20, 20, 12], start=1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    rr = 2
    for order in orders:
        computed = sum((it.qty or 0) * (it.unit_price or 0) for it in order.items)
        row_vals = [order.source_file, order.order_no, order.order_type,
                    order.buyer, order.order_date, len(order.items),
                    round(computed, 2), order.stated_total, None]
        for c, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=rr, column=c, value=val)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c in (1, 3, 4) else CENTER
        ws2.cell(row=rr, column=7).number_format = "#,##0.00"
        ws2.cell(row=rr, column=8).number_format = "#,##0.00"
        ws2.cell(row=rr, column=9, value=f'=IF(ABS(G{rr}-H{rr})<0.01,"OK","FARK")')
        rr += 1
    ws2.freeze_panes = "A2"
    wb.save(out_path)


def process_files(pdf_paths: List[str], out_path: str):
    """PDF listesini işler, (orders, log_satirlari) döner."""
    orders, log = [], []
    for pdf in pdf_paths:
        try:
            o = parse_order(pdf)
            orders.append(o)
            log.append(f"✓ {os.path.basename(pdf)} — {len(o.items)} kalem "
                       f"(sipariş {o.order_no or '?'})")
        except Exception as e:
            log.append(f"✗ {os.path.basename(pdf)} — HATA: {e}")
    if orders:
        build_workbook(orders, out_path)
    return orders, log


# ============================================================================
#  3) GRAFİK ARAYÜZ (tkinter)  —  tembel import, Windows'ta çalışır
# ============================================================================

def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    class App(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Liebherr PO → Excel  |  CEKAPIM")
            self.geometry("640x520")
            self.minsize(560, 460)
            self.files: List[str] = []
            self._build()

        def _build(self):
            pad = {"padx": 10, "pady": 6}
            top = tk.Frame(self)
            top.pack(fill="x", **pad)
            tk.Label(top, text="Liebherr Sipariş PDF'leri → Excel",
                     font=("Segoe UI", 13, "bold")).pack(anchor="w")
            tk.Label(top, fg="#555",
                     text="PDF ekle, çıktı Excel'i seç, Dönüştür'e bas. "
                          "Tek veya çok kalemli PO'lar desteklenir.").pack(anchor="w")

            btns = tk.Frame(self)
            btns.pack(fill="x", **pad)
            tk.Button(btns, text="＋ PDF Ekle", width=14,
                      command=self.add_files).pack(side="left")
            tk.Button(btns, text="Seçileni Sil", width=12,
                      command=self.remove_selected).pack(side="left", padx=6)
            tk.Button(btns, text="Listeyi Temizle", width=13,
                      command=self.clear_files).pack(side="left")

            frame = tk.Frame(self)
            frame.pack(fill="both", expand=True, padx=10)
            tk.Label(frame, text="Seçilen dosyalar:").pack(anchor="w")
            self.listbox = tk.Listbox(frame, selectmode="extended", height=8)
            self.listbox.pack(fill="both", expand=True)

            out = tk.Frame(self)
            out.pack(fill="x", **pad)
            tk.Label(out, text="Çıktı:").pack(side="left")
            self.out_var = tk.StringVar(value=os.path.join(
                os.path.expanduser("~"), "Desktop", "liebherr_siparisler.xlsx"))
            tk.Entry(out, textvariable=self.out_var).pack(
                side="left", fill="x", expand=True, padx=6)
            tk.Button(out, text="...", width=3,
                      command=self.choose_output).pack(side="left")

            self.run_btn = tk.Button(self, text="⬇  Excel'e Dönüştür",
                                     bg="#1F4E78", fg="white",
                                     font=("Segoe UI", 11, "bold"),
                                     command=self.convert)
            self.run_btn.pack(fill="x", padx=10, pady=(4, 6))

            self.log = tk.Text(self, height=7, bg="#F5F5F5", state="disabled")
            self.log.pack(fill="both", expand=False, padx=10, pady=(0, 10))

        # --- olaylar --------------------------------------------------------
        def add_files(self):
            paths = filedialog.askopenfilenames(
                title="Liebherr PO PDF'lerini seç",
                filetypes=[("PDF dosyaları", "*.pdf"), ("Tüm dosyalar", "*.*")])
            for p in paths:
                if p not in self.files:
                    self.files.append(p)
                    self.listbox.insert("end", os.path.basename(p))

        def remove_selected(self):
            for i in reversed(self.listbox.curselection()):
                self.listbox.delete(i)
                del self.files[i]

        def clear_files(self):
            self.listbox.delete(0, "end")
            self.files.clear()

        def choose_output(self):
            p = filedialog.asksaveasfilename(
                title="Excel'i kaydet", defaultextension=".xlsx",
                initialfile="liebherr_siparisler.xlsx",
                filetypes=[("Excel", "*.xlsx")])
            if p:
                self.out_var.set(p)

        def _log(self, msg):
            self.log.configure(state="normal")
            self.log.insert("end", msg + "\n")
            self.log.see("end")
            self.log.configure(state="disabled")

        def convert(self):
            if not self.files:
                messagebox.showwarning("Uyarı", "Önce en az bir PDF ekleyin.")
                return
            out = self.out_var.get().strip()
            if not out:
                messagebox.showwarning("Uyarı", "Çıktı dosyası seçin.")
                return
            self.run_btn.configure(state="disabled", text="İşleniyor…")
            threading.Thread(target=self._worker, args=(list(self.files), out),
                             daemon=True).start()

        def _worker(self, files, out):
            try:
                orders, log = process_files(files, out)
                for line in log:
                    self.after(0, self._log, line)
                if orders:
                    total = sum(len(o.items) for o in orders)
                    self.after(0, self._log,
                               f"\n✓ Bitti: {len(orders)} sipariş, {total} kalem.")
                    self.after(0, self._log, f"→ {out}")
                    self.after(0, lambda: messagebox.showinfo(
                        "Tamamlandı",
                        f"{len(orders)} sipariş, {total} kalem işlendi.\n\n{out}"))
                else:
                    self.after(0, lambda: messagebox.showerror(
                        "Hata", "Hiçbir PDF ayrıştırılamadı."))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Hata", str(e)))
            finally:
                self.after(0, lambda: self.run_btn.configure(
                    state="normal", text="⬇  Excel'e Dönüştür"))

    App().mainloop()


# ============================================================================
#  4) Giriş noktası — GUI aç; komut satırı argümanı verilirse CLI çalışır
# ============================================================================

def run_cli(args):
    pdfs = []
    out = "liebherr_siparisler.xlsx"
    i = 0
    while i < len(args):
        if args[i] in ("-o", "--output"):
            out = args[i + 1]; i += 2; continue
        p = args[i]
        if os.path.isdir(p):
            pdfs += sorted(glob.glob(os.path.join(p, "*.pdf")))
        elif p.lower().endswith(".pdf"):
            pdfs.append(p)
        i += 1
    if not pdfs:
        print("PDF bulunamadı."); sys.exit(1)
    orders, log = process_files(pdfs, out)
    for line in log:
        print(line)
    print(f"\n→ {out}" if orders else "Ayrıştırılamadı.")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        run_cli(sys.argv[1:])
    else:
        run_gui()